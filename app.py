"""
==============================================================================
API - Consulta de Aulas por Docente
Universidad Privada Domingo Savio (UPDS)
==============================================================================

DESCRIPCIÓN GENERAL:
    Este archivo contiene la API Flask que gestiona las consultas de aulas
    para docentes. Incluye funcionalidades de:
    - Consulta pública para docentes
    - Administración (CRUD) de registros
    - Carga de archivos Excel para actualizar la base de datos

ENDPOINTS PÚBLICOS (Consulta):
    GET  /                    → Información general de la API
    GET  /api/health          → Verificación de estado del servicio
    GET  /api/docentes        → Lista todos los docentes disponibles
    GET  /api/consulta        → Consulta aulas por nombre de docente
    GET  /api/aulas           → Lista todas las asignaciones con filtros

ENDPOINTS ADMINISTRATIVOS (CRUD):
    GET    /api/admin/registros     → Lista todos los registros con paginación
    POST   /api/admin/registros     → Crear nuevo registro
    PUT    /api/admin/registros/<id>→ Actualizar registro existente
    DELETE /api/admin/registros/<id>→ Eliminar registro
    POST   /api/admin/upload        → Subir archivo Excel para reemplazar datos
    GET    /api/admin/export        → Exportar datos actuales a JSON

ESTRUCTURA DE UN REGISTRO:
    {
        "id": 1,                    # Identificador único (autogenerado)
        "turno": "MAÑANA",          # MAÑANA | MEDIO DIA | TARDE | NOCHE
        "materia": "CALCULO I",     # Nombre de la materia
        "docente": "Juan Pérez",    # Nombre completo del docente
        "aula": "A-101",            # Código del aula
        "horario": "08:00 - 09:30"  # Horario de la clase
    }

DEPENDENCIAS:
    - Flask >= 3.0.0
    - openpyxl (para leer archivos Excel)

AUTOR: Sistema de Consulta de Aulas UPDS
VERSIÓN: 2.0 (con módulo de administración)
==============================================================================
"""

# =============================================================================
# IMPORTACIONES
# =============================================================================
from flask import Flask, jsonify, request, send_file
import json
import os
import unicodedata
from datetime import datetime
import io

# Intentamos importar openpyxl para la carga de Excel
# Si no está disponible, la funcionalidad de upload se deshabilitará
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("ADVERTENCIA: openpyxl no está instalado. La carga de Excel estará deshabilitada.")

# =============================================================================
# CONFIGURACIÓN DE LA APLICACIÓN FLASK
# =============================================================================
app = Flask(__name__)

# Configuración para subida de archivos
# Limitamos el tamaño máximo a 16MB para evitar problemas de memoria
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo


# =============================================================================
# CARGA Y GESTIÓN DE DATOS
# =============================================================================
"""
Los datos se almacenan en un archivo JSON (consultas_data.json).
Este archivo se carga al iniciar la aplicación y se guarda cada vez
que hay modificaciones (crear, actualizar, eliminar registros).

La estructura del archivo es una lista de diccionarios, donde cada
diccionario representa una asignación de aula.
"""

# Ruta al archivo de datos JSON
DATA_PATH = os.path.join(os.path.dirname(__file__), "consultas_data.json")

# Variable global que contiene todos los registros en memoria
ASIGNACIONES: list[dict] = []

def cargar_datos():
    """
    Carga los datos desde el archivo JSON al iniciar la aplicación.
    
    Si el archivo no existe, inicializa una lista vacía.
    Esta función se ejecuta automáticamente al importar el módulo.
    
    Returns:
        None (modifica la variable global ASIGNACIONES)
    """
    global ASIGNACIONES
    
    if os.path.exists(DATA_PATH):
        with open(DATA_PATH, "r", encoding="utf-8") as f:
            datos = json.load(f)
            
            # Aseguramos que cada registro tenga un ID único
            # Esto es necesario para la compatibilidad con datos antiguos
            for i, registro in enumerate(datos):
                if "id" not in registro:
                    registro["id"] = i + 1
            
            ASIGNACIONES = datos
            print(f"✓ Datos cargados: {len(ASIGNACIONES)} registros")
    else:
        ASIGNACIONES = []
        print("⚠ Archivo de datos no encontrado. Iniciando con lista vacía.")


def guardar_datos():
    """
    Guarda los datos actuales al archivo JSON.
    
    Esta función se llama automáticamente después de cualquier
    operación que modifique los datos (crear, actualizar, eliminar).
    
    Returns:
        bool: True si se guardó correctamente, False en caso de error
    """
    try:
        with open(DATA_PATH, "w", encoding="utf-8") as f:
            json.dump(ASIGNACIONES, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"ERROR al guardar datos: {e}")
        return False


def generar_nuevo_id():
    """
    Genera un nuevo ID único para un registro.
    
    El ID es el máximo ID existente + 1, o 1 si no hay registros.
    
    Returns:
        int: Nuevo ID único
    """
    if not ASIGNACIONES:
        return 1
    
    max_id = max(r.get("id", 0) for r in ASIGNACIONES)
    return max_id + 1


# Cargamos los datos al iniciar el módulo
cargar_datos()


# =============================================================================
# FUNCIONES UTILITARIAS
# =============================================================================

def normalizar(texto: str) -> str:
    """
    Normaliza texto para búsquedas: minúsculas, sin acentos, sin espacios extra.
    
    Esta función es crucial para que las búsquedas funcionen correctamente
    sin importar cómo el usuario escriba el nombre (con o sin acentos,
    mayúsculas o minúsculas).
    
    Ejemplo:
        normalizar("José María") → "jose maria"
        normalizar("CÁLCULO") → "calculo"
    
    Args:
        texto (str): Texto a normalizar
    
    Returns:
        str: Texto normalizado
    """
    # Paso 1: Descomponer caracteres Unicode (ej: é → e + ́)
    nfkd = unicodedata.normalize("NFKD", texto)
    
    # Paso 2: Eliminar marcas diacríticas (acentos)
    sin_acentos = "".join(c for c in nfkd if not unicodedata.combining(c))
    
    # Paso 3: Convertir a minúsculas y eliminar espacios extra
    return sin_acentos.lower().strip()


def buscar_docente(query: str) -> list[dict]:
    """
    Busca todas las asignaciones que coinciden con el query del docente.
    
    La búsqueda es parcial y no sensible a acentos ni mayúsculas.
    
    Ejemplo:
        buscar_docente("Miranda") 
        → Encuentra "Miranda Hoyos Victor", "Juan Miranda", etc.
    
    Args:
        query (str): Texto a buscar en el nombre del docente
    
    Returns:
        list[dict]: Lista de asignaciones que coinciden
    """
    q = normalizar(query)
    resultados = []
    
    for item in ASIGNACIONES:
        docente_norm = normalizar(item.get("docente", ""))
        if q in docente_norm:
            resultados.append(item)
    
    return resultados


def buscar_materia(query: str) -> list[dict]:
    """
    Busca todas las asignaciones que coinciden con el query de materia.
    
    La búsqueda es parcial y no sensible a acentos ni mayúsculas.
    
    Ejemplo:
        buscar_materia("Calculo") 
        → Encuentra "CÁLCULO I", "CÁLCULO II", "PRE-CÁLCULO", etc.
    
    Args:
        query (str): Texto a buscar en el nombre de la materia
    
    Returns:
        list[dict]: Lista de asignaciones que coinciden
    """
    q = normalizar(query)
    resultados = []
    
    for item in ASIGNACIONES:
        materia_norm = normalizar(item.get("materia", ""))
        if q in materia_norm:
            resultados.append(item)
    
    return resultados


def get_docentes_unicos() -> list[str]:
    """
    Obtiene la lista de docentes únicos, excluyendo 'NO DEFINIDO'.
    
    Returns:
        list[str]: Lista ordenada alfabéticamente de nombres de docentes
    """
    docs = set()
    
    for item in ASIGNACIONES:
        nombre = item.get("docente", "").strip()
        if nombre and nombre.upper() != "NO DEFINIDO":
            docs.add(nombre)
    
    # Ordenamos alfabéticamente usando la versión normalizada para comparar
    return sorted(docs, key=lambda x: normalizar(x))


def get_materias_unicas() -> list[str]:
    """
    Obtiene la lista de materias únicas.
    
    Returns:
        list[str]: Lista ordenada alfabéticamente de nombres de materias
    """
    materias = set()
    
    for item in ASIGNACIONES:
        materia = item.get("materia", "").strip()
        if materia:
            materias.add(materia)
    
    # Ordenamos alfabéticamente usando la versión normalizada para comparar
    return sorted(materias, key=lambda x: normalizar(x))


def validar_registro(datos: dict, es_actualizacion: bool = False) -> tuple[bool, str]:
    """
    Valida que un registro tenga todos los campos requeridos.
    
    Args:
        datos (dict): Diccionario con los datos del registro
        es_actualizacion (bool): Si es True, los campos vacíos se permiten
                                 (se mantienen los valores anteriores)
    
    Returns:
        tuple: (es_valido: bool, mensaje_error: str)
    """
    campos_requeridos = ["turno", "materia", "docente", "aula", "horario"]
    
    if not es_actualizacion:
        # Para nuevos registros, todos los campos son obligatorios
        for campo in campos_requeridos:
            if campo not in datos or not str(datos[campo]).strip():
                return False, f"El campo '{campo}' es obligatorio"
    
    # Validar que el turno sea válido si se proporciona
    turnos_validos = ["MAÑANA", "MEDIO DIA", "TARDE", "NOCHE"]
    if "turno" in datos and datos["turno"]:
        turno_upper = datos["turno"].upper().strip()
        if turno_upper not in turnos_validos:
            return False, f"Turno inválido. Valores permitidos: {', '.join(turnos_validos)}"
    
    return True, ""


# =============================================================================
# ENDPOINTS PÚBLICOS - CONSULTA
# =============================================================================

@app.route("/", methods=["GET"])
def raiz():
    """
    Endpoint raíz - Muestra información general de la API.
    
    Este endpoint sirve como documentación básica para desarrolladores
    y para verificar que la API está funcionando.
    
    Returns:
        JSON con información de la API y lista de endpoints disponibles
    """
    return jsonify({
        "message": "API de Consulta de Aulas - UPDS",
        "version": "2.0",
        "docs": "https://github.com/bolivianotech/consulta-aulas-upds",
        "endpoints": {
            "/api/health": "Verificación de estado del servicio",
            "/api/docentes": "Lista todos los docentes (opcional: ?q=filtro)",
            "/api/consulta": "Consulta aulas por docente (?docente=nombre&turno=opcional)",
            "/api/aulas": "Lista todas las asignaciones con filtros opcionales",
            "/api/admin/registros": "CRUD de registros (GET, POST, PUT, DELETE)",
            "/api/admin/upload": "Subir archivo Excel para actualizar datos"
        }
    })


@app.route("/api/health", methods=["GET"])
def health():
    """
    Verificación de estado del servicio (health check).
    
    Este endpoint es utilizado por servicios de monitoreo y por
    plataformas como Render para verificar que la API está funcionando.
    
    Returns:
        JSON con estado y estadísticas básicas
    """
    return jsonify({
        "status": "ok",
        "total_asignaciones": len(ASIGNACIONES),
        "total_docentes": len(get_docentes_unicos()),
        "excel_support": EXCEL_SUPPORT,
        "timestamp": datetime.now().isoformat()
    })


@app.route("/api/docentes", methods=["GET"])
def docentes():
    """
    Lista todos los docentes disponibles.
    
    Este endpoint se usa principalmente para el autocompletado
    en la interfaz de búsqueda.
    
    Query params opcionales:
        ?q=<texto>   → Filtro por nombre (búsqueda parcial)
    
    Ejemplos:
        GET /api/docentes           → Todos los docentes
        GET /api/docentes?q=Mir     → Docentes que contienen "Mir"
    
    Returns:
        JSON con lista de docentes y total
    """
    q = request.args.get("q", "").strip()
    
    if q:
        # Filtrar docentes que coincidan con el query
        q_norm = normalizar(q)
        docs = [
            d for d in get_docentes_unicos()
            if q_norm in normalizar(d)
        ]
    else:
        docs = get_docentes_unicos()

    return jsonify({
        "docentes": docs,
        "total": len(docs)
    })


@app.route("/api/sugerencias", methods=["GET"])
def sugerencias():
    """
    Endpoint de autocompletado combinado: busca tanto en docentes como en materias.
    
    Este endpoint es usado por la interfaz de búsqueda para mostrar
    sugerencias mientras el usuario escribe.
    
    Query params:
        ?q=<texto>   → Texto a buscar (mínimo 2 caracteres)
    
    Ejemplos:
        GET /api/sugerencias?q=Mir     → Docentes y materias que contienen "Mir"
        GET /api/sugerencias?q=calc    → Encontrará "CÁLCULO I", "CÁLCULO II", etc.
    
    Returns:
        JSON con lista de sugerencias (tipo: 'docente' o 'materia')
    """
    q = request.args.get("q", "").strip()
    
    if len(q) < 2:
        return jsonify({"sugerencias": [], "total": 0})
    
    q_norm = normalizar(q)
    sugerencias_list = []
    
    # Buscar en docentes
    for docente in get_docentes_unicos():
        if q_norm in normalizar(docente):
            sugerencias_list.append({
                "texto": docente,
                "tipo": "docente"
            })
    
    # Buscar en materias (únicas)
    materias_unicas = get_materias_unicas()
    for materia in materias_unicas:
        if q_norm in normalizar(materia):
            sugerencias_list.append({
                "texto": materia,
                "tipo": "materia"
            })
    
    # Limitar a 10 sugerencias (5 docentes + 5 materias máximo)
    docentes_sug = [s for s in sugerencias_list if s["tipo"] == "docente"][:5]
    materias_sug = [s for s in sugerencias_list if s["tipo"] == "materia"][:5]
    sugerencias_list = docentes_sug + materias_sug
    
    return jsonify({
        "sugerencias": sugerencias_list,
        "total": len(sugerencias_list)
    })


@app.route("/api/consulta", methods=["GET"])
def consulta():
    """
    Consulta principal: busca aulas por docente O por materia.
    
    Este endpoint permite buscar tanto por nombre de docente como por
    nombre de materia. Primero intenta buscar por docente, y si no
    encuentra resultados, busca por materia.
    
    Query params:
        ?q=<texto>          (OBLIGATORIO) Texto a buscar (docente o materia)
        ?docente=<nombre>   (alternativo) Búsqueda solo por docente
        ?turno=<turno>      (opcional)    Filtrar por turno
    
    Turnos válidos:
        MAÑANA, MEDIO DIA, TARDE, NOCHE
    
    Ejemplos:
        GET /api/consulta?q=Miranda           (busca docente "Miranda")
        GET /api/consulta?q=Calculo           (busca materia "Cálculo")
        GET /api/consulta?docente=Miranda     (compatibilidad anterior)
        GET /api/consulta?q=Miranda&turno=NOCHE
    
    Returns:
        JSON con información de la búsqueda y sus asignaciones
    """
    # Soportamos tanto 'q' (nuevo) como 'docente' (compatibilidad)
    query = request.args.get("q", "").strip() or request.args.get("docente", "").strip()
    turno_filtro = request.args.get("turno", "").strip().upper()

    # Validación: se requiere algún término de búsqueda
    if not query:
        return jsonify({
            "error": "Parámetro 'q' o 'docente' es requerido.",
            "ejemplo": "/api/consulta?q=Miranda o /api/consulta?q=Calculo"
        }), 400

    # Primero buscar por docente
    resultados = buscar_docente(query)
    tipo_busqueda = "docente"
    
    # Si no hay resultados por docente, buscar por materia
    if not resultados:
        resultados = buscar_materia(query)
        tipo_busqueda = "materia"

    # Aplicar filtro por turno si se proporcionó
    if turno_filtro:
        resultados = [r for r in resultados if r.get("turno", "").upper() == turno_filtro]

    # Determinar el valor encontrado (docente o materia)
    if resultados:
        if tipo_busqueda == "docente":
            encontrado = resultados[0]["docente"]
        else:
            encontrado = resultados[0]["materia"]
    else:
        encontrado = None

    return jsonify({
        "tipo_busqueda": tipo_busqueda,
        "encontrado": encontrado,
        "consulta": query,
        "turno_filtro": turno_filtro or None,
        "total_asignaciones": len(resultados),
        "asignaciones": resultados
    })


@app.route("/api/aulas", methods=["GET"])
def aulas():
    """
    Lista todas las asignaciones con filtros opcionales.
    
    Este endpoint permite explorar todas las asignaciones
    aplicando diferentes filtros.
    
    Query params opcionales:
        ?turno=<turno>       Filtrar por turno
        ?materia=<nombre>    Filtrar por materia (búsqueda parcial)
        ?aula=<codigo>       Filtrar por código de aula (búsqueda parcial)
    
    Ejemplos:
        GET /api/aulas                    → Todas las asignaciones
        GET /api/aulas?turno=MAÑANA       → Solo turno mañana
        GET /api/aulas?materia=calculo    → Materias que contienen "calculo"
    
    Returns:
        JSON con filtros aplicados y lista de asignaciones
    """
    turno = request.args.get("turno", "").strip().upper()
    materia = request.args.get("materia", "").strip()
    aula = request.args.get("aula", "").strip()

    resultados = ASIGNACIONES.copy()

    # Aplicar filtros
    if turno:
        resultados = [r for r in resultados if r.get("turno", "").upper() == turno]
    
    if materia:
        mat_norm = normalizar(materia)
        resultados = [r for r in resultados if mat_norm in normalizar(r.get("materia", ""))]
    
    if aula:
        aula_norm = normalizar(aula)
        resultados = [r for r in resultados if aula_norm in normalizar(r.get("aula", ""))]

    return jsonify({
        "total": len(resultados),
        "filtros": {
            "turno": turno or None,
            "materia": materia or None,
            "aula": aula or None
        },
        "asignaciones": resultados
    })


# =============================================================================
# ENDPOINTS ADMINISTRATIVOS - CRUD
# =============================================================================

@app.route("/api/admin/registros", methods=["GET"])
def admin_listar_registros():
    """
    Lista todos los registros con paginación y búsqueda.
    
    Este endpoint es usado por el panel de administración para
    mostrar la tabla de registros.
    
    Query params opcionales:
        ?page=<numero>       Página actual (default: 1)
        ?per_page=<numero>   Registros por página (default: 20, max: 100)
        ?search=<texto>      Búsqueda en todos los campos
        ?turno=<turno>       Filtrar por turno
    
    Returns:
        JSON con registros paginados y metadatos de paginación
    """
    # Obtener parámetros de paginación
    try:
        page = max(1, int(request.args.get("page", 1)))
        per_page = min(100, max(1, int(request.args.get("per_page", 20))))
    except ValueError:
        page = 1
        per_page = 20
    
    # Obtener parámetros de filtrado
    search = request.args.get("search", "").strip()
    turno = request.args.get("turno", "").strip().upper()
    
    # Filtrar registros
    resultados = ASIGNACIONES.copy()
    
    if turno:
        resultados = [r for r in resultados if r.get("turno", "").upper() == turno]
    
    if search:
        search_norm = normalizar(search)
        resultados = [
            r for r in resultados
            if (search_norm in normalizar(r.get("docente", "")) or
                search_norm in normalizar(r.get("materia", "")) or
                search_norm in normalizar(r.get("aula", "")) or
                search_norm in normalizar(r.get("horario", "")))
        ]
    
    # Calcular paginación
    total = len(resultados)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    
    start = (page - 1) * per_page
    end = start + per_page
    
    return jsonify({
        "registros": resultados[start:end],
        "paginacion": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        },
        "filtros": {
            "search": search or None,
            "turno": turno or None
        }
    })


@app.route("/api/admin/registros", methods=["POST"])
def admin_crear_registro():
    """
    Crea un nuevo registro de asignación.
    
    Body (JSON):
        {
            "turno": "MAÑANA",
            "materia": "CALCULO I",
            "docente": "Juan Pérez",
            "aula": "A-101",
            "horario": "08:00 - 09:30"
        }
    
    Todos los campos son obligatorios.
    
    Returns:
        JSON con el registro creado (incluyendo el ID asignado)
    """
    # Verificar que se envió JSON
    if not request.is_json:
        return jsonify({
            "error": "Se requiere Content-Type: application/json"
        }), 400
    
    datos = request.get_json()
    
    # Validar datos
    es_valido, mensaje_error = validar_registro(datos)
    if not es_valido:
        return jsonify({"error": mensaje_error}), 400
    
    # Crear nuevo registro
    nuevo_registro = {
        "id": generar_nuevo_id(),
        "turno": datos["turno"].upper().strip(),
        "materia": datos["materia"].strip(),
        "docente": datos["docente"].strip(),
        "aula": datos["aula"].strip(),
        "horario": datos["horario"].strip()
    }
    
    # Agregar a la lista y guardar
    ASIGNACIONES.append(nuevo_registro)
    guardar_datos()
    
    return jsonify({
        "mensaje": "Registro creado exitosamente",
        "registro": nuevo_registro
    }), 201


@app.route("/api/admin/registros/<int:registro_id>", methods=["GET"])
def admin_obtener_registro(registro_id):
    """
    Obtiene un registro específico por su ID.
    
    Args:
        registro_id (int): ID del registro a obtener
    
    Returns:
        JSON con el registro encontrado o error 404
    """
    registro = next((r for r in ASIGNACIONES if r.get("id") == registro_id), None)
    
    if not registro:
        return jsonify({"error": f"Registro con ID {registro_id} no encontrado"}), 404
    
    return jsonify({"registro": registro})


@app.route("/api/admin/registros/<int:registro_id>", methods=["PUT"])
def admin_actualizar_registro(registro_id):
    """
    Actualiza un registro existente.
    
    Solo se actualizan los campos que se envían en el body.
    Los campos no enviados mantienen su valor anterior.
    
    Args:
        registro_id (int): ID del registro a actualizar
    
    Body (JSON):
        {
            "turno": "TARDE",      // opcional
            "materia": "...",      // opcional
            "docente": "...",      // opcional
            "aula": "...",         // opcional
            "horario": "..."       // opcional
        }
    
    Returns:
        JSON con el registro actualizado
    """
    # Buscar el registro
    registro = next((r for r in ASIGNACIONES if r.get("id") == registro_id), None)
    
    if not registro:
        return jsonify({"error": f"Registro con ID {registro_id} no encontrado"}), 404
    
    # Verificar que se envió JSON
    if not request.is_json:
        return jsonify({"error": "Se requiere Content-Type: application/json"}), 400
    
    datos = request.get_json()
    
    # Validar datos (permitiendo campos vacíos en actualizaciones)
    es_valido, mensaje_error = validar_registro(datos, es_actualizacion=True)
    if not es_valido:
        return jsonify({"error": mensaje_error}), 400
    
    # Actualizar solo los campos proporcionados
    campos_actualizables = ["turno", "materia", "docente", "aula", "horario"]
    
    for campo in campos_actualizables:
        if campo in datos and datos[campo] is not None:
            valor = str(datos[campo]).strip()
            if valor:  # Solo actualizar si el valor no está vacío
                if campo == "turno":
                    valor = valor.upper()
                registro[campo] = valor
    
    # Guardar cambios
    guardar_datos()
    
    return jsonify({
        "mensaje": "Registro actualizado exitosamente",
        "registro": registro
    })


@app.route("/api/admin/registros/<int:registro_id>", methods=["DELETE"])
def admin_eliminar_registro(registro_id):
    """
    Elimina un registro por su ID.
    
    Args:
        registro_id (int): ID del registro a eliminar
    
    Returns:
        JSON con mensaje de confirmación o error 404
    """
    global ASIGNACIONES
    
    # Buscar el índice del registro
    indice = next((i for i, r in enumerate(ASIGNACIONES) if r.get("id") == registro_id), None)
    
    if indice is None:
        return jsonify({"error": f"Registro con ID {registro_id} no encontrado"}), 404
    
    # Guardar el registro para el mensaje de respuesta
    registro_eliminado = ASIGNACIONES[indice]
    
    # Eliminar el registro
    ASIGNACIONES.pop(indice)
    guardar_datos()
    
    return jsonify({
        "mensaje": "Registro eliminado exitosamente",
        "registro_eliminado": registro_eliminado
    })


# =============================================================================
# ENDPOINTS ADMINISTRATIVOS - CARGA DE EXCEL
# =============================================================================

@app.route("/api/admin/upload", methods=["POST"])
def admin_upload_excel():
    """
    Sube un archivo Excel para reemplazar todos los datos.
    
    El archivo debe ser un .xlsx con la siguiente estructura:
    - Fila 1: Puede contener título (se ignora)
    - Fila 2: Headers (NRO, TURNO, MATERIA, DOCENTE, AULA, HORARIO)
    - Fila 3+: Datos
    
    Las columnas esperadas son (desde columna B):
        B: NRO (número de fila, se ignora)
        C: TURNO
        D: MATERIA
        E: DOCENTE
        F: AULA
        G: HORARIO
    
    IMPORTANTE: Esta operación REEMPLAZA todos los datos existentes.
    
    Returns:
        JSON con estadísticas de la carga
    """
    global ASIGNACIONES
    
    # Verificar que openpyxl esté disponible
    if not EXCEL_SUPPORT:
        return jsonify({
            "error": "La carga de Excel no está disponible. Instale openpyxl: pip install openpyxl"
        }), 500
    
    # Verificar que se envió un archivo
    if "file" not in request.files:
        return jsonify({"error": "No se envió ningún archivo. Use el campo 'file'"}), 400
    
    archivo = request.files["file"]
    
    # Verificar que el archivo tiene nombre
    if archivo.filename == "":
        return jsonify({"error": "El archivo no tiene nombre"}), 400
    
    # Verificar extensión
    if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({
            "error": "Solo se permiten archivos Excel (.xlsx, .xls)"
        }), 400
    
    try:
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        
        nuevos_registros = []
        errores = []
        fila_num = 0
        
        # Procesar filas desde la fila 3 (datos)
        # Columnas: B=NRO, C=TURNO, D=MATERIA, E=DOCENTE, F=AULA, G=HORARIO
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=7, values_only=True):
            fila_num += 1
            nro, turno, materia, docente, aula, horario = row
            
            # Saltar filas vacías (sin número)
            if nro is None:
                continue
            
            # Validar datos mínimos
            if not turno or not materia:
                errores.append(f"Fila {fila_num + 2}: Faltan datos obligatorios (turno o materia)")
                continue
            
            # Crear registro
            registro = {
                "id": len(nuevos_registros) + 1,
                "turno": str(turno).strip().upper() if turno else "",
                "materia": str(materia).strip() if materia else "",
                "docente": str(docente).strip() if docente else "NO DEFINIDO",
                "aula": str(aula).strip() if aula else "",
                "horario": str(horario).strip() if horario else ""
            }
            
            nuevos_registros.append(registro)
        
        # Verificar que se encontraron registros
        if not nuevos_registros:
            return jsonify({
                "error": "No se encontraron registros válidos en el archivo",
                "errores": errores
            }), 400
        
        # Reemplazar datos existentes
        registros_anteriores = len(ASIGNACIONES)
        ASIGNACIONES = nuevos_registros
        guardar_datos()
        
        return jsonify({
            "mensaje": "Archivo procesado exitosamente",
            "estadisticas": {
                "registros_anteriores": registros_anteriores,
                "registros_nuevos": len(nuevos_registros),
                "docentes_unicos": len(get_docentes_unicos()),
                "errores_encontrados": len(errores)
            },
            "errores": errores if errores else None
        })
        
    except Exception as e:
        return jsonify({
            "error": f"Error al procesar el archivo: {str(e)}"
        }), 500


@app.route("/api/admin/export", methods=["GET"])
def admin_exportar_json():
    """
    Exporta todos los datos actuales como archivo JSON.
    
    Útil para hacer respaldos de los datos antes de cargar
    un nuevo archivo Excel.
    
    Returns:
        Archivo JSON descargable
    """
    # Crear el contenido JSON
    contenido = json.dumps(ASIGNACIONES, ensure_ascii=False, indent=2)
    
    # Generar nombre con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"consultas_backup_{timestamp}.json"
    
    # Crear archivo en memoria
    buffer = io.BytesIO()
    buffer.write(contenido.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/json',
        as_attachment=True,
        download_name=nombre_archivo
    )


# =============================================================================
# CONFIGURACIÓN CORS
# =============================================================================
"""
CORS (Cross-Origin Resource Sharing) permite que la landing page
(que puede estar en un dominio diferente, como GitHub Pages)
pueda hacer peticiones a esta API (que está en Render).

Sin esto, el navegador bloquearía las peticiones por seguridad.
"""

@app.after_request
def add_cors_headers(response):
    """
    Agrega headers CORS a todas las respuestas.
    
    Esto permite que cualquier origen (*) pueda acceder a la API.
    En un entorno de producción real, podrías restringir esto
    a dominios específicos.
    """
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================

if __name__ == "__main__":
    """
    Inicia el servidor de desarrollo Flask.
    
    En producción (Render), Gunicorn o similar manejará esto.
    El servidor escucha en todas las interfaces (0.0.0.0) en el puerto 5000.
    """
    print("\n" + "=" * 60)
    print("API de Consulta de Aulas - UPDS")
    print("=" * 60)
    print(f"Total registros cargados: {len(ASIGNACIONES)}")
    print(f"Total docentes únicos: {len(get_docentes_unicos())}")
    print(f"Soporte Excel: {'Sí' if EXCEL_SUPPORT else 'No'}")
    print("=" * 60)
    print("\nEndpoints disponibles:")
    print("  GET  /                         → Info de la API")
    print("  GET  /api/health               → Health check")
    print("  GET  /api/docentes             → Lista docentes")
    print("  GET  /api/consulta?docente=... → Consulta aulas")
    print("  GET  /api/admin/registros      → Lista registros (admin)")
    print("  POST /api/admin/upload         → Subir Excel (admin)")
    print("=" * 60 + "\n")
    
    app.run(debug=True, host="0.0.0.0", port=5000)
